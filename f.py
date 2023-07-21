from  tkinter import *
import tkinter as tk 
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk

root = tk.Tk()
root.geometry("1120x600")
root.title("Login Interface")
root.resizable(False, False)

excel_con = Workbook()
excel_con = load_workbook("Sample_file.xlsx")
excel_data = excel_con['Data']
excel_old = excel_con["Old Students"]
excel_new = excel_con["New Students"]

def oldReg():
	newroot = Toplevel()
	newroot.geometry("600x500")
	newroot.title("Register Interface For Old Students") 
	newroot.resizable(False, False)

	path = "l.webp"
	bg = Image.open(path)
	resize_bg = bg.resize((600, 500))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(newroot, height=500, width=600)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(270, 70, text="Register For Old Students", font=("Arial", 20),fill="black")

	studentnol = LabelFrame(newroot, text="Student No.", width=15, bg="white")
	studentno = Entry(studentnol,font=("Arial", 18))
	studentno.pack()
	studentno_w = my_canva.create_window(130, 100, anchor="nw", window=studentnol)

	year_sectionl = LabelFrame(newroot, text="Year And Section", width=15, bg="white")
	year_section = Entry(year_sectionl,font=("Arial", 18))
	year_section.pack()
	year_section_w = my_canva.create_window(130, 170, anchor="nw", window=year_sectionl)

	passwordl = LabelFrame(newroot, text="Password", width=15, bg="white")
	password = Entry(passwordl,font=("Arial", 18))
	password.pack()
	password_w = my_canva.create_window(130, 240, anchor="nw", window=passwordl)

	def oldlyreg():
		excel_old.append((studentno.get(), password.get(), year_section.get()))
		messagebox.showinfo("Register Message", "Register Successful")
		excel_con.save("Sample_file.xlsx")

	r_button = Button(newroot, text="Register", width=20, bg="lightblue", font=("Arial", 12) , command=lambda:oldlyreg())
	button = my_canva.create_window(170, 310, anchor="nw", window=r_button)

	newroot.mainloop()
	

def newReg():
	
	nroot = Toplevel()
	nroot.geometry("600x500")
	nroot.title("Register Interface For New Students") 
	nroot.resizable(False, False)

	path = "r.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((600, 500))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(nroot, height=500, width=600)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(260, 60, text="Register For New Students", font=("Arial", 20),fill="black")

	fullnl = LabelFrame(nroot, text="Fullname", width=15, bg="white")
	fulln = Entry(fullnl,font=("Arial", 18))
	fulln.pack()
	fulln_w = my_canva.create_window(130, 90, anchor="nw", window=fullnl)

	year_secl = LabelFrame(nroot, text="Year And Section", width=15, bg="white")
	year_sec = Entry(year_secl,font=("Arial", 18))
	year_sec.pack()
	year_sec_w = my_canva.create_window(130, 160, anchor="nw", window=year_secl)

	passwl = LabelFrame(nroot, text="Password", width=15, bg="white")
	passw = Entry(passwl,font=("Arial", 18))
	passw.pack()
	passw_w = my_canva.create_window(130, 230, anchor="nw", window=passwl)

	def newlyreg():
		excel_new.append((fulln.get(), passw.get(), year_sec.get()))
		messagebox.showinfo("Register Message", "Register Successful")
		excel_con.save("Sample_file.xlsx")

	r_button1 = Button(nroot, text="Register", width=20, bg="lightblue", font=("Arial",12),command=lambda:newlyreg())
	r_button = my_canva.create_window(170, 300, anchor="nw", window=r_button1)

	nroot.mainloop()

def reg():
	if str_var.get() == "Old":
		oldReg()
	elif str_var.get() == "New":
		newReg()
	else:
		messagebox.showerror("REGISTER ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")
		
def newLog():
	groot = Toplevel()
	groot.geometry("700x350")
	groot.title("Login Interface For New Students") 
	#groot.resizable(False, False)

	path = "login.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((700, 350))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(groot, height=350, width=700)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(450, 50, text="Sign In For New Students", font=("Arial", 20),fill="black")

	fullnamel = LabelFrame(groot, text="Fullname", width=15, bg="white")
	fullname = Entry(fullnamel,font=("Arial", 18))
	fullname.pack()
	fullname_w = my_canva.create_window(310, 90, anchor="nw", window=fullnamel)

	passwordLabel = LabelFrame(groot, text="Password", width=15, bg="white")
	passwor = Entry(passwordLabel,font=("Arial", 18), show="*")
	passwor.pack()
	password = my_canva.create_window(310, 170, anchor="nw", window=passwordLabel)

	def getUsers(fullname, passwor):
		id = 1
		isExisted = False
		for data in excel_new.iter_rows(values_only=True):
			if data[0] == fullname and data[1] == passwor:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			groot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			groot.destroy()

	button1 = Button(groot, text="Login", width=20, bg="lightblue", font=("Arial", 12), command=lambda:getUsers(fullname.get(), passwor.get()))
	button1 = my_canva.create_window(350, 250, anchor="nw", window=button1)
	groot.mainloop()	


def oldLog():
	vroot = Toplevel()
	vroot.geometry("700x350")
	vroot.title("Login Interface For Old Students") 
	#vroot.resizable(False, False)

	path = "login.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((700, 350))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(vroot, height=350, width=700)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(450, 50, text="Sign In For Old Students", font=("Arial", 20),fill="black")

	studentnl = LabelFrame(vroot, text="Student No", width=15, bg="white")
	studentn = Entry(studentnl,font=("Arial", 18))
	studentn.pack()
	studentn_w = my_canva.create_window(310, 90, anchor="nw", window=studentnl)

	passwol = LabelFrame(vroot, text="Password", width=15, bg="white")
	passwo = Entry(passwol,font=("Arial", 18), show="*")
	passwo.pack()
	passwo_w = my_canva.create_window(310, 170, anchor="nw", window=passwol)

	def getUser(studentn, passwo):
		id = 1
		isExisted = False
		for data in excel_old.iter_rows(values_only=True):
			if data[0] == studentn and data[1] == passwo:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			vroot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			vroot.destroy()

	button1 = Button(vroot, text="Login", width=20, bg="lightblue", font=("Arial",12), command=lambda:getUser(studentn.get(), passwo.get()))
	button1 = my_canva.create_window(350, 250, anchor="nw", window=button1)

	vroot.mainloop()	
def log():
	if str_var.get() == "Old":
		oldLog()
	elif str_var.get() == "New":
		newLog()
	else:
		messagebox.showerror("LOGIN ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")


	


#FRAMES
topframe = Frame(root, height=50, width=1120, bg="#023047")
topframe.pack(fill="x", side="top")
leftframe = Frame(root, height=600, width=350, bg="#023047", bd=0)
leftframe.pack(fill="y",side='left')
rigthframe = Frame(root, height=600, width=820)
rigthframe.pack(fill="both", side='right')

#TEXT IN TOPFRAME
l = Label(topframe,text="Dalubhasaan Ng Lungsod Ng Lucena", bg="#023047", fg="#FFB703", font=("Arial", 30), justify="center")
l.pack()

#LOGO IMAGE
b = (Image.open("dll_logo.png"))
resized_image = b.resize((190, 190))
new_image = ImageTk.PhotoImage(resized_image)
logo = Label(leftframe, image=new_image, bg="#023047")
logo.pack()

#RadioButton
rFrame = Frame(leftframe)

str_var = StringVar()
str_var.set("none")
old_R = Radiobutton(rFrame, text="Old", variable=str_var, value="Old", font=("Arial", 15),bg="#023047",fg="#FFB703")
new_R = Radiobutton(rFrame, text="New", variable=str_var, value="New", font=("Arial", 15),bg="#023047",fg="#FFB703")
old_R.grid(row=0, column=0)
new_R.grid(row=0, column=1)

rFrame.pack()

#Buttons
rb = Button(leftframe, text="Register", width=15, font=("Arial", 12), command=lambda: reg())
lb = Button(leftframe, text="Login", width=15, font=("Arial", 12), command=lambda: log())
rb.pack(pady=15)
lb.pack()

#TEXT BUTTON 
def showProf():
	prof.pack(fill="both")
	abt.pack_forget()

def showAbt():
	abt.pack(fill="both")
	prof.pack_forget()

abt_b = Button(topframe, text="About", command=showAbt, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
abt_b.pack(side=RIGHT, anchor=N)
prof_b = Button(topframe, text="Profile", command=showProf, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
prof_b.pack(side=RIGHT, anchor=N)

#FRAME FOR THE PROFILE AND ABOUT
prof = Frame(rigthframe, bg="#023047")

#SCHOOL IMAGE
bg = (Image.open("school.jpg"))
resized_imag = bg.resize((1120, 600))
new_imag = ImageTk.PhotoImage(resized_imag)
my_canva = Canvas(prof, height=600, width=1080)
my_canva.pack()
my_canva.create_image(0,0,image=new_imag, anchor="nw")
prof.pack()

#About
a = "Dalubhasaan ng Lungsod ng Lucena"
b = "One of the high impact programs of Mayor Roderick A. Alcala is free quality tertiary education."
c = "When he assumed office in 2012, Dalubhasaan ng Lungsod ng Lucena (DLL) was his vision of providing access to college education for free. \nMayor Alcala envisions DLL as an institution that would provide easy access to higher education and \nultimately develop the competencies of the youth of the city to meet the demands of the local industries and businesses."
d = "Through DLL, students from low-income families are able to enrol in degree programs at no cost. \nThe annual appropriation of the local government has allowed DLL to cover all its operation expenses \nincluding tuition and miscellaneous fees of students. \nThe college, operated, managed, and fully-subsidized by the City Government, implements a zero-collection policy."
e = "At present, DLL has a total of nine degree programs and continue to apply for additional academic programs to accommodate more scholars:"
f = "Bachelor of Arts in Information Technology"
g = "Bachelor of Arts in Public Administration"
h = "Bachelor of Science in Accountancy"
i = "Bachelor of Science in Accounting Information System"
j = "Bachelor of Science in English Language Studies"
k = "Bachelor of Science in Entrepreneurship"
l = "Bachelor of Science in Social Work"
m = "Bachelor in Technical Vocational Teacher’s Education"
n = "Diploma in Hotel and Restaurant Services"
o = "At present the former annex building of Lucena City Hall is being renovated to accommodate the growing student population of DLL."

abt = Frame(rigthframe, width=600, height=750)
a_l = Label(abt,text=a,font=("System" ,25),fg="black",)
b_l = Label(abt,text=b,font=("System" ,11),fg="black",)
c_l = Label(abt,text=c,font=("System" ,11),fg="black",)
d_l = Label(abt,text=d,font=("System" ,11),fg="black",)
e_l = Label(abt,text=e,font=("System" ,11),fg="black",)
f_l = Label(abt,text=f,font=("System" ,11),fg="black",)
g_l = Label(abt,text=g,font=("System" ,11),fg="black",)
h_l = Label(abt,text=h,font=("System" ,11),fg="black",)
i_l = Label(abt,text=i,font=("System" ,11),fg="black",)
j_l = Label(abt,text=j,font=("System" ,11),fg="black",)
k_l = Label(abt,text=k,font=("System" ,11),fg="black",)
l_l = Label(abt,text=l,font=("System" ,11),fg="black",)
m_l = Label(abt,text=m,font=("System" ,11),fg="black",)
n_l = Label(abt,text=n,font=("System" ,11),fg="black",)
o_l = Label(abt,text=o,font=("System" ,11),fg="black",)

a_l.pack()
b_l.pack()
c_l.pack()
d_l.pack()
e_l.pack()
f_l.pack()
g_l.pack()
h_l.pack()
i_l.pack()
j_l.pack()
k_l.pack()
l_l.pack()
m_l.pack()
n_l.pack()
o_l.pack()


abt.pack()
showProf()


root.mainloop()

