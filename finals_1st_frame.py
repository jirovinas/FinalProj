from  tkinter import *
import tkinter as tk 
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk
from finals_2nd_frame import newinterface 

#FIRST INTERFACE
root = tk.Tk()
root.geometry("1120x600")
root.title("Login Interface")
root.resizable(False, False)

excel_con = Workbook()
excel_con = load_workbook("Sample_file.xlsx")
excel_data = excel_con['Data']
excel_old = excel_con["Old Students"]
excel_new = excel_con["New Students"]


def oldReg():
	newroot = Toplevel()
	newroot.geometry("600x500")
	newroot.title("Register Interface For Old Students") 
	newroot.resizable(False, False)

	path = "l.webp"
	bg = Image.open(path)
	resize_bg = bg.resize((600, 500))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(newroot, height=500, width=600)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(270, 70, text="Register For Old Students", font=("Arial", 20),fill="black")

	studentnol = LabelFrame(newroot, text="Student No.", width=15, bg="white")
	studentno = Entry(studentnol,font=("Arial", 18))
	studentno.pack()
	studentno_w = my_canva.create_window(130, 100, anchor="nw", window=studentnol)

	year_sectionl = LabelFrame(newroot, text="Year And Section", width=15, bg="white")
	year_section = Entry(year_sectionl,font=("Arial", 18))
	year_section.pack()
	year_section_w = my_canva.create_window(130, 170, anchor="nw", window=year_sectionl)

	passwordl = LabelFrame(newroot, text="Password", width=15, bg="white")
	password = Entry(passwordl,font=("Arial", 18))
	password.pack()
	password_w = my_canva.create_window(130, 240, anchor="nw", window=passwordl)

	def oldlyreg():
		s = studentno.get()
		ys = year_section.get()
		ps = password.get()
		if s == "":
			messagebox.showerror("Error", "Please Enter Your Student No.")
		elif ys == "":
			messagebox.showerror("Error", "Please Enter Your Year and Section ")
		elif ps == "":
			messagebox.showerror("Error", "Please Enter Your Password")
		else:
			excel_old.append((studentno.get(), password.get(), year_section.get()))
			messagebox.showinfo("Register Message", "Register Successful")
			excel_con.save("Sample_file.xlsx")
			newroot.withdraw()

	r_button = Button(newroot, text="Register", width=20, bg="lightblue", font=("Arial", 12) , command=lambda:oldlyreg())
	button = my_canva.create_window(170, 310, anchor="nw", window=r_button)

	newroot.mainloop()
	

def newReg():
	
	nroot = Toplevel()
	nroot.geometry("600x500")
	nroot.title("Register Interface For New Students") 
	nroot.resizable(False, False)

	path = "r.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((600, 500))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(nroot, height=500, width=600)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(260, 60, text="Register For New Students", font=("Arial", 20),fill="black")

	fullnl = LabelFrame(nroot, text="Fullname", width=15, bg="white")
	fulln = Entry(fullnl,font=("Arial", 18))
	fulln.pack()
	fulln_w = my_canva.create_window(130, 90, anchor="nw", window=fullnl)

	year_secl = LabelFrame(nroot, text="Year And Section", width=15, bg="white")
	year_sec = Entry(year_secl,font=("Arial", 18))
	year_sec.pack()
	year_sec_w = my_canva.create_window(130, 160, anchor="nw", window=year_secl)

	passwl = LabelFrame(nroot, text="Password", width=15, bg="white")
	passw = Entry(passwl,font=("Arial", 18))
	passw.pack()
	passw_w = my_canva.create_window(130, 230, anchor="nw", window=passwl)

	def newlyreg():
		f = fulln.get()
		p = passw.get()
		y = year_sec.get()
		if f == "":
			messagebox.showerror("Error", "Please Enter Your Fullname")
		elif p == "":
			messagebox.showerror("Error", "Please Enter Your Fullname")
		elif y == "":
			messagebox.showerror("Error", "Please Enter Your Year and Section")
		else:
			excel_new.append((fulln.get(), passw.get(), year_sec.get()))
			messagebox.showinfo("Register Message", "Register Successful")
			excel_con.save("Sample_file.xlsx")
			nroot.withdraw()

	r_button1 = Button(nroot, text="Register", width=20, bg="lightblue", font=("Arial",12),command=lambda:newlyreg())
	r_button = my_canva.create_window(170, 300, anchor="nw", window=r_button1)

	nroot.mainloop()

#Register Function
def reg():
	if str_var.get() == "Old":
		oldReg()
	elif str_var.get() == "New":
		newReg()
	else:
		messagebox.showerror("REGISTER ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")
		
def newLog():
	groot = Toplevel()
	groot.geometry("700x350")
	groot.title("Login Interface For New Students") 
	#groot.resizable(False, False)

	path = "login.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((700, 350))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(groot, height=350, width=700)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(450, 50, text="Sign In For New Students", font=("Arial", 20),fill="black")

	fullnamel = LabelFrame(groot, text="Fullname", width=15, bg="white")
	fullname = Entry(fullnamel,font=("Arial", 18))
	fullname.pack()
	fullname_w = my_canva.create_window(310, 90, anchor="nw", window=fullnamel)

	passwordLabel = LabelFrame(groot, text="Password", width=15, bg="white")
	passwor = Entry(passwordLabel,font=("Arial", 18), show="*")
	passwor.pack()
	password = my_canva.create_window(310, 170, anchor="nw", window=passwordLabel)

	def getUsers(fullname, passwor):
		id = 1
		isExisted = False
		for data in excel_new.iter_rows(values_only=True):
			if data[0] == fullname and data[1] == passwor:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			groot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			groot.destroy()
			root.withdraw()
			newinterface()

	button1 = Button(groot, text="Login", width=20, bg="lightblue", font=("Arial", 12), command=lambda:getUsers(fullname.get(), passwor.get()))
	button1 = my_canva.create_window(350, 250, anchor="nw", window=button1)
	groot.mainloop()	


def oldLog():
	vroot = Toplevel()
	vroot.geometry("700x350")
	vroot.title("Login Interface For Old Students") 
	#vroot.resizable(False, False)

	path = "login.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((700, 350))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(vroot, height=350, width=700)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(450, 50, text="Sign In For Old Students", font=("Arial", 20),fill="black")

	studentnl = LabelFrame(vroot, text="Student No", width=15, bg="white")
	studentn = Entry(studentnl,font=("Arial", 18))
	studentn.pack()
	studentn_w = my_canva.create_window(310, 90, anchor="nw", window=studentnl)

	passwol = LabelFrame(vroot, text="Password", width=15, bg="white")
	passwo = Entry(passwol,font=("Arial", 18), show="*")
	passwo.pack()
	passwo_w = my_canva.create_window(310, 170, anchor="nw", window=passwol)

	def getUser(studentn, passwo):
		id = 1
		isExisted = False
		for data in excel_old.iter_rows(values_only=True):
			if data[0] == studentn and data[1] == passwo:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			vroot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			vroot.destroy()
			root.withdraw()
			newinterface()

	button1 = Button(vroot, text="Login", width=20, bg="lightblue", font=("Arial",12), command=lambda:getUser(studentn.get(), passwo.get()))
	button1 = my_canva.create_window(350, 250, anchor="nw", window=button1)

	vroot.mainloop()
#Login function	
def log():
	if str_var.get() == "Old":
		oldLog()
	elif str_var.get() == "New":
		newLog()
	else:
		messagebox.showerror("LOGIN ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")


#FRAMES
topframe = Frame(root, height=50, width=1120, bg="#023047")
topframe.pack(fill="x", side="top")
leftframe = Frame(root, height=600, width=350, bg="#023047", bd=0)
leftframe.pack(fill="y",side='left')
rigthframe = Frame(root, height=600, width=820)
rigthframe.pack(fill="both", side='right')

#TEXT IN TOPFRAME
l = Label(topframe,text="Dalubhasaan Ng Lungsod Ng Lucena", bg="#023047", fg="#FFB703", font=("Arial", 30), justify="center")
l.pack()

#LOGO IMAGE
b = (Image.open("dll_logo.png"))
resized_image = b.resize((190, 190))
new_image = ImageTk.PhotoImage(resized_image)
logo = Label(leftframe, image=new_image, bg="#023047")
logo.pack()

#RadioButton
rFrame = Frame(leftframe)

str_var = StringVar()
str_var.set("none")
old_R = Radiobutton(rFrame, text="Old", variable=str_var, value="Old", font=("Arial", 15),bg="#023047",fg="#FFB703")
new_R = Radiobutton(rFrame, text="New", variable=str_var, value="New", font=("Arial", 15),bg="#023047",fg="#FFB703")
old_R.grid(row=0, column=0)
new_R.grid(row=0, column=1)

rFrame.pack()

#Buttons
rb = Button(leftframe, text="Register", width=15, font=("Arial", 12), command=lambda: reg())
lb = Button(leftframe, text="Login", width=15, font=("Arial", 12), command=lambda: log())
rb.pack(pady=15)
lb.pack()

def animateAbout():
	global aboutIterate, txt
	if aboutIterate < len(a):
		txt += a[aboutIterate] + "\n"
		a_l1.config(text=txt)
		aboutIterate += 1
		a_l1.after(200, animateAbout)

#TEXT BUTTON 
def showProf():
	prof.pack(fill="both")
	abt.pack_forget()

def showAbt():
	abt.pack(fill="both")
	prof.pack_forget()
	animateAbout()

abt_b = Button(topframe, text="About", command=showAbt, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
abt_b.pack(side=RIGHT, anchor=N)
prof_b = Button(topframe, text="Profile", command=showProf, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
prof_b.pack(side=RIGHT, anchor=N)

#FRAME FOR THE PROFILE AND ABOUT
prof = Frame(rigthframe, bg="#023047")

#SCHOOL IMAGE
bg = (Image.open("school.jpg"))
resized_imag = bg.resize((1120, 600))
new_imag = ImageTk.PhotoImage(resized_imag)
my_canva = Canvas(prof, height=600, width=1080)
my_canva.pack()
my_canva.create_image(0,0,image=new_imag, anchor="nw")
prof.pack()

aboutIterate = 0
txt = ""

#About
_a = "Dalubhasaan ng Lungsod ng Lucena"
a = [
	"One of the high impact programs of Mayor Roderick A. Alcala is free quality tertiary education.",
	"When he assumed office in 2012, Dalubhasaan ng Lungsod ng Lucena (DLL) was his vision of providing access to college education for free. \nMayor Alcala envisions DLL as an institution that would provide easy access to higher education and \nultimately develop the competencies of the youth of the city to meet the demands of the local industries and businesses.",
	"Through DLL, students from low-income families are able to enrol in degree programs at no cost. \nThe annual appropriation of the local government has allowed DLL to cover all its operation expenses \nincluding tuition and miscellaneous fees of students. \nThe college, operated, managed, and fully-subsidized by the City Government, implements a zero-collection policy.",
	"At present, DLL has a total of nine degree programs and continue to apply for additional academic programs to accommodate more scholars:",
	"Bachelor of Arts in Information Technology",
	"Bachelor of Arts in Public Administration",
	"Bachelor of Science in Accountancy",
	"Bachelor of Science in Accounting Information System",
	"Bachelor of Science in English Language Studies",
	"Bachelor of Science in Entrepreneurship",
	"Bachelor of Science in Social Work",
	"Bachelor in Technical Vocational Teacher’s Education",
	"Diploma in Hotel and Restaurant Services",
	"At present the former annex building of Lucena City Hall is being renovated to accommodate the growing student population of DLL."
]

abt = Frame(rigthframe, width=600, height=750)
a_l = Label(abt,text=_a,font=("System" ,25),fg="black",justify="center")
a_l1 = Label(abt,font=("System" ,16),fg="black",justify="left")

a_l.pack()
a_l1.pack(pady=20)


abt.pack()
abt.mainloop()

showProf()


root.mainloop()

