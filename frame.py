from cgitb import text
from http.client import FOUND
from re import search
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk

#New Interface
NewRoot = tk.Tk()
NewRoot.geometry("1030x820")
NewRoot.title("New Interface")
NewRoot.configure(bg="black")

#Excel File
excel_con = Workbook()
excel_con = load_workbook("Sample_file.xlsx")
excel_activate = excel_con.active

#Frames
topframe = Frame(NewRoot, height=70, width=1030, bg="#CCAAFF")
topframe.grid(row=0, column=0, columnspan=6, rowspan=2, sticky=W)
leftframe = Frame(NewRoot, height=600, width=1030, bd=3)
leftframe.grid(row=2, column=0, columnspan=3, rowspan=6, sticky=W)
bottomframe = Frame(NewRoot, height=100, width=1205, bg="#CCAAFF")
bottomframe.grid(row=8, column=0, columnspan=6, rowspan=2, sticky=W)

# topframe
pat = "gg.jpg"
photo = ImageTk.PhotoImage(file = pat)
my_canv = Canvas(topframe, height=70, width=1030)
my_canv.grid()
my_canv.create_image(0,0,image=photo, anchor="nw")
my_canv.create_text(515, 40, text="STUDENT RECORDS SYSTEM", font=("Arial", 40),fill="white")

# leftframe
path = "gg.jpg"
photo0 = ImageTk.PhotoImage(file = path)
my_canva = Canvas(leftframe, height=600, width=1030)
my_canva.grid(row=0, column=0)
my_canva.create_image(0,0,image=photo0, anchor="nw")

student_no = LabelFrame(leftframe, text="Student No.")
studentno_Entry = Entry(student_no, width=50)
student_no.grid(row=0, column=0)
studentno_Entry.grid(row=0, column=1, padx=10, pady=10,)
student_no_canva = my_canva.create_window(50, 20, anchor="nw", window=student_no)

full_name = LabelFrame(leftframe, text="Full Name")
fullname_Entry = Entry(full_name, width=50)
full_name.grid(row=1, column=0)
fullname_Entry.grid(row=1, column=1, padx=10, pady=10)
full_name_canva = my_canva.create_window(50, 90, anchor="nw", window=full_name)

email = LabelFrame(leftframe, text="Email")
email_Entry = Entry(email, width=50)
email.grid(row=2, column=0)
email_Entry.grid(row=2, column=1, padx=10, pady=10)
email_canva = my_canva.create_window(50, 160, anchor="nw", window=email)

gender = LabelFrame(leftframe, text="Gender")
gender_var = StringVar()
male_R = Radiobutton(gender, text="Male", variable=gender_var, value="Male", width=20)
female_R = Radiobutton(gender, text="Female", variable=gender_var, value="Female", width=18)
male_R.grid(row=3, column=0)
female_R.grid(row=3, column=1)
gender_canva = my_canva.create_window(50, 230, anchor="nw", window=gender)

course = LabelFrame(leftframe, text="Course")
course_var = StringVar()
course_list = ["BSIT", "BSA", "BSAIS", "ABELS", "BSSW", "BSE", "DHRS", "BSPA", "BTVTE"]
course_combo = ttk.Combobox(course, values=course_list, textvariable=course_var,font=("Times New Roman", 20))
course_combo.grid(row=4, column=0)
course_canva = my_canva.create_window(50, 300, anchor="nw", window=course)

no = LabelFrame(leftframe, text="Contact No.")
no_Entry = Entry(no, width=50)
no.grid(row=5, column=0)
no_Entry.grid(row=5, column=1, padx=10, pady=10)
no_canva = my_canva.create_window(50, 370, anchor="nw", window=no)

address = LabelFrame(leftframe, text="Address")
address_Entry = Entry(address, width=50)
address.grid(row=6, column=0)
address_Entry.grid(row=7, column=1, padx=10, pady=10)
address_canva = my_canva.create_window(50, 440, anchor="nw", window=address)

# rightframe
student_details = LabelFrame(leftframe, text="Student Details")
student_detailsT = Text(student_details, width=44, font=("", 18), height=18)
student_details.grid(row=1, column=1, columnspan=2)
student_detailsT.grid(row=2, column=1)
student_details_canva = my_canva.create_window(420, 40, anchor="nw", window=student_details)

# bottomframe
pa = "gg.jpg"
pho = ImageTk.PhotoImage(file = pa)
my_can = Canvas(bottomframe, height=150, width=1030)
my_can.grid(row=0, column=0)
my_can.create_image(0,0,image=pho, anchor="nw")

# r_button = Button(newroot, text="Register", width=20, bg="lightblue", command=lambda:oldlyreg())
# button = my_canva.create_window(170, 230, anchor="nw", window=r_button)

add_btn = Button(bottomframe, text="Add Record", font=("System", 20))
add_btn_w = my_can.create_window(5, 5, anchor="nw", window=add_btn)
# add_btn.grid(row=0, column=0)

save_btn = Button(bottomframe, text="Save", font=("System", 20))
save_btn_w = my_can.create_window(5, 70, anchor="nw", window=save_btn)
# save_btn = Button(bottomframe, text="Save", command=save_records, width=15, pady=5, font=("", 20), bg="#CCBBFF")
# save_btn.grid(row=0, column=1)

search_btn = Button(bottomframe, text="Search", font=("System", 20))
search_btn_w = my_can.create_window(200, 5, anchor="nw", window=search_btn)
# search_btn = Button(bottomframe, text="Search", command=search_data, width=15, pady=5, font=("", 20), bg="#CCBBFF")
# search_btn.grid(row=1, column=0)

view_btn = Button(bottomframe, text="View Data", font=("System", 20))
view_btn_w = my_can.create_window(200, 70, anchor="nw", window=view_btn)
# view_btn=Button(bottomframe,text="View Data", command=view_data ,width=15, pady=5, font=("", 20), bg="#CCBBFF")
# view_btn.grid(row=1, column=1)


print_btn = Button(bottomframe, text="Print", font=("System", 20))
print_btn_w = my_can.create_window(500, 40, anchor="nw", window=print_btn)
# print_btn = Button(bottomframe, text="Print", command=print_records, width=15, pady=5, font=("", 20), bg="#CCBBFF")
# print_btn.grid(row=0, column=2, rowspan=2)

reset_btn = Button(bottomframe, text="Reset", font=("System", 20))
reset_btn_w = my_can.create_window(700, 5, anchor="nw", window=reset_btn)
# reset_btn = Button(bottomframe, text="Reset", command=reset_fields, width=15, pady=5, font=("", 20), bg="#CCBBFF")
# reset_btn.grid(row=0, column=3)   

exit_btn = Button(bottomframe, text="Exit", font=("System", 20))
exit_btn_w = my_can.create_window(700, 70, anchor="nw", window=exit_btn)
# exit_btn = Button(bottomframe, text="Exit", command=exit_interface, width=15, pady=5, font=("", 20), bg="#CCBBFF")
# exit_btn.grid(row=0, column=4)



edit_btn = Button(bottomframe, text="Edit", font=("System", 20))
edit_btn_w = my_can.create_window(950, 5, anchor="nw", window=edit_btn)
# edit_btn = Button(bottomframe, text="Edit", command=edit_data ,width=15, pady=5, font=("", 20), bg="#CCBBFF")
# edit_btn.grid(row=1, column=4)

del_btn = Button(bottomframe, text="Delete", font=("System", 20))
del_btn_w = my_can.create_window(910, 70, anchor="nw", window=del_btn)
# del_btn=Button(bottomframe,text="Delete", command=delete_data ,width=15, pady=5, font=("", 20), bg="#CCBBFF")
# del_btn.grid(row=1, column=3)

NewRoot.mainloop()