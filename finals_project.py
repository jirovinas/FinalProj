from cgitb import text
from http.client import FOUND
from re import search
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk

def login():
    if username.get() == "" and password.get() == "":
        messagebox.showinfo("Login", "Login Successfully")
        root.withdraw()
        newInterface()
    else:
        messagebox.showerror("Login", "Login Denied")

def newInterface():
    def add_record():
        student_no = studentno_Entry.get()
        full_name = fullname_Entry.get()
        email = email_Entry.get()
        gender = gender_Entry.get()
        course = course_Entry.get()
        contact_no = no_Entry.get()
        allowance = allowance_Entry.get()
        address = address_Entry.get()
        
        record = f"Student Ref: {student_no}\n"
        record += f"Full Name: {full_name}\n"
        record += f"Email: {email}\n"
        record += f"Gender: {gender}\n"
        record += f"course: {course}\n"
        record += f"Contact No.: {contact_no}\n"
        record += f"allowance: {allowance}\n"
        record += f"Address: {address}\n"
        
        student_detailsT.insert(END, record)

        studentno_Entry.delete(0, END)
        fullname_Entry.delete(0, END)
        email_Entry.delete(0, END)
        gender_Entry.delete(0, END)
        course_Entry.delete(0, END)
        no_Entry.delete(0, END)
        allowance_Entry.delete(0, END)
        address_Entry.delete(0, END)
    
    def print_records():
        student_details_text = student_detailsT.get("1.0", END)
        messagebox.showinfo("Print Records", student_details_text)
        student_details_text = student_detailsT.delete("1.0", END)

    def reset_fields():
        studentno_Entry.delete(0, END)
        fullname_Entry.delete(0, END)
        email_Entry.delete(0, END)
        gender_Entry.delete(0, END)
        course_Entry.delete(0, END)
        no_Entry.delete(0, END)
        allowance_Entry.delete(0, END)
        address_Entry.delete(0, END)
        student_detailsT.delete("1.0", END)
    
    def exit_interface():
        NewRoot.destroy()
    
    NewRoot = tk.Toplevel()
    NewRoot.geometry("1200x760")
    NewRoot.title("New Interface")
    NewRoot.configure(bg="black")
    excel_con = Workbook()
    excel_con = load_workbook("Sample_file.xlsx")
    excel_activate = excel_con.active
    
    # def search_data():
    #     searchTl = tk.Toplevel()
    #     searchTl.geometry("300x130")
    #     searchTl.title("SEARCH")
        
    #     def search_name():
    #         Found = False
    #         search_n = search_E.get()
    #         for each_cell in range(2, excel_activate.max_row + 1):
    #             if search_n == excel_activate["A"+ str(each_cell)].value:
    #                 cell_address = str(each_cell)
    #                 Found = True
    #                 break
    #         if Found == True:
    #             messagebox.showinfo("Data", f"Data Found @ Record Number {cell_address}")
    #             searchTl.destroy()
    #         else:
    #             messagebox.showerror("Data", "Data Not Found")

    #     def exit_search():
    #         searchTl.destroy()
    #     search_L = Label(searchTl, text="What Is Your Student No.", font=("Arial", 18), width=20)
    #     search_E = Entry(searchTl, width=20, font=("Arial", 20))
    #     search_B = Button(searchTl, text="SEARCH", command=lambda:search_name(), width=20, bg="#b1f2ff")
    #     search_Ex = Button(searchTl, text="EXIT", command=lambda:exit_search(), width=20, bg="#b1f2ff")

    #     search_L.grid(row=0, sticky=W)
    #     search_E.grid(row=1, sticky=W)
    #     search_B.grid(row=2)
    #     search_Ex.grid(row=3)

    #     searchTl.mainloop()
    def search_data():
        NewRoot = tk.Toplevel()
        NewRoot.geometry("600x300")
        NewRoot.title("Search Data")
        excel_con = Workbook()
        excel_con = load_workbook("Sample_file.xlsx")
        excel_activate = excel_con.active
        data = []

        for i in excel_activate.iter_rows(values_only=True):
            data.append(i)

        def exit_search():
            NewRoot.destroy()

        def search_name():
            for i in tree.get_children():
                tree.delete(i)
            lists = []
            for i in data:
                x = False
                for j in i:
                    if j != None:
                        if search_box.get().lower() in j.lower():
                            x = True
                            break
                if x:
                    lists.append(i)
            
            for i in lists:
                tree.insert('', index=END, values=i)


        search_box = Entry(NewRoot, width=30)
        search_box.pack()

        search_button = Button(NewRoot, text='Search', width=10, command=search_name)
        search_button.pack()

        exit_button = Button(NewRoot, text='Exit', width=10, command=exit_search)
        exit_button.pack()

        tree = ttk.Treeview(NewRoot, show='headings')
        treescrolly = Scrollbar(NewRoot, orient="vertical", command=tree.yview)
        treescrollx = Scrollbar(NewRoot, orient="horizontal", command=tree.xview)
        tree.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
        treescrollx.pack(side ="bottom",fill ="x")
        treescrolly.pack(side ="right",fill="y")  
        columns=("Student No.", "Fullname", "Email", "Gender", "Course", "Contact No.", "Allowance", "Address")

        tree['columns'] = columns
        for i in columns:
            tree.heading(i, text=i)
            tree.column(i)

        for i, row in enumerate(data):
            tree.insert('', END, text=str(i), values=row)

        tree.pack()

        NewRoot.mainloop()

    def save_records():
        student_no = studentno_Entry.get()
        full_name = fullname_Entry.get()
        email = email_Entry.get()
        gender = gender_Entry.get()
        course = course_Entry.get()
        contact_no = no_Entry.get()
        allowance = allowance_Entry.get()
        address = address_Entry.get()

        Found = False
        for each_cell in range(2, excel_activate.max_row + 1):
            if student_no == excel_activate["A"+ str(each_cell)].value or full_name == excel_activate["B"+ str(each_cell)].value or email ==  excel_activate['C'+str(each_cell)].value or gender ==  excel_activate['D'+str(each_cell)].value or course ==  excel_activate['E'+str(each_cell)].value or contact_no ==  excel_activate['F'+str(each_cell)].value or allowance ==  excel_activate['G'+str(each_cell)].value or address ==  excel_activate['H'+str(each_cell)].value:
                Found = True
                break
        if Found == True:
            messagebox.showerror("Data", "Data Already Exist")
        else:
            lastrow = str(excel_activate.max_row + 1)
            excel_activate["A"+lastrow] = student_no
            excel_activate["B"+lastrow] = full_name
            excel_activate["C"+lastrow] = email
            excel_activate["D"+lastrow] = gender
            excel_activate["E"+lastrow] = course
            excel_activate["F"+lastrow] = contact_no
            excel_activate["G"+lastrow] = allowance
            excel_activate["H"+lastrow] = address

            excel_con.save("Sample_file.xlsx")
            messagebox.showinfo("Save Records", "Records saved successfully!")
            refresh_data(tv1)

    def delete_data():
        student_no = studentno_Entry.get()
        full_name = fullname_Entry.get()
        email = email_Entry.get()
        gender = gender_Entry.get()
        course = course_Entry.get()
        contact_no = no_Entry.get()
        allowance = allowance_Entry.get()
        address = address_Entry.get()
        for each_cell in range(2, (excel_activate.max_row)+1):
            if student_no ==  excel_activate['A'+str(each_cell)].value or full_name ==  excel_activate['B'+str(each_cell)].value or email ==  excel_activate['C'+str(each_cell)].value or gender ==  excel_activate['D'+str(each_cell)].value or course ==  excel_activate['E'+str(each_cell)].value or contact_no ==  excel_activate['F'+str(each_cell)].value or allowance ==  excel_activate['G'+str(each_cell)].value or address ==  excel_activate['H'+str(each_cell)].value: 
                Found = True
                cell_address = each_cell
                break
            else:
                Found=False
        if(Found == True):
            excel_activate.delete_rows(cell_address)
            messagebox.showinfo("INFO","DATA DELETED")
            clear_entries()
        excel_con.save('Sample_file.xlsx')
        refresh_data(tv1)

    def clear_entries():
        studentno_Entry.delete(0, END)
        fullname_Entry.delete(0, END)
        email_Entry.delete(0, END)
        gender_Entry.delete(0, END)
        course_Entry.delete(0, END)
        no_Entry.delete(0, END)
        allowance_Entry.delete(0, END)
        address_Entry.delete(0, END)
    
    def refresh_data(tree):
        tree.delete(*tree.get_children())
        data = get_updated_data()
        for each_cell in range(2, (excel_activate.max_row)+1):
            tv1.insert(parent='', index="end", text=str(each_cell),values=(excel_activate['A'+str(each_cell)].value,excel_activate['B'+str(each_cell)].value, excel_activate['C'+str(each_cell)].value, excel_activate['D'+str(each_cell)].value, excel_activate['E'+str(each_cell)].value, excel_activate['F'+str(each_cell)].value, excel_activate['G'+str(each_cell)].value, excel_activate['H'+str(each_cell)].value))
        
    def get_updated_data():
        updated_value = list()
        for each_cell in range(2, (excel_activate.max_row)+1):     
            updated_value.append([excel_activate['A'+str(each_cell)].value,excel_activate['B'+str(each_cell)].value, excel_activate['C'+str(each_cell)].value, excel_activate['D'+str(each_cell)].value,excel_activate['E'+str(each_cell)].value,excel_activate['F'+str(each_cell)].value,excel_activate['G'+str(each_cell)].value,excel_activate['H'+str(each_cell)].value])
        return updated_value

    def view_data():
        view_frame = Toplevel()
        view_frame.geometry('600x300')
        view_frame.title('View  from Excel')
 
        global tv1
        tv1 = ttk.Treeview(view_frame)
        treescrolly = Scrollbar(view_frame, orient="vertical", command=tv1.yview)
        treescrollx = Scrollbar(view_frame, orient="horizontal", command=tv1.xview)
        tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
        treescrollx.pack(side ="bottom",fill ="x")
        treescrolly.pack(side ="right",fill="y")  

        tv1['columns'] = ("Student No.", "Fullname", "Email", "Gender", "Course", "Contact No.", "Allowance", "Address")
        tv1.column("#0", width=120, minwidth=25)
        tv1.column("Student No.", anchor=W, width=120)
        tv1.column("Fullname",  anchor=W, width=120)
        tv1.column("Email", anchor=W, width=120)
        tv1.column("Gender", anchor=W, width=120)
        tv1.column("Course", anchor=W, width=120)
        tv1.column("Contact No.", anchor=W, width=120)
        tv1.column("Allowance", anchor=W, width=120)
        tv1.column("Address", anchor=W, width=120)

        tv1.heading("#0", text="Label", anchor=W)
        tv1.heading("Student No.", text="Student No.", anchor=W)
        tv1.heading("Fullname", text="Fullname", anchor=W)
        tv1.heading("Email", text="Email", anchor=W)
        tv1.heading("Gender", text="Gender", anchor=W)
        tv1.heading("Course", text="Course", anchor=W)
        tv1.heading("Contact No.", text="Contact No.", anchor=W)
        tv1.heading("Allowance", text="Allowance", anchor=W)
        tv1.heading("Address", text="Address", anchor=W)


        for each_cell in range(2, (excel_activate.max_row)+1):
            tv1.insert(parent='', index="end", text=str(each_cell),values=(excel_activate['A'+str(each_cell)].value,excel_activate['B'+str(each_cell)].value, excel_activate['C'+str(each_cell)].value, excel_activate['D'+str(each_cell)].value, excel_activate['E'+str(each_cell)].value, excel_activate['F'+str(each_cell)].value, excel_activate['G'+str(each_cell)].value, excel_activate['H'+str(each_cell)].value))
        tv1.pack()
        view_frame.mainloop()
        refresh_data(tv1)

    

    def edit_data():
        student_no = studentno_Entry.get()
        for each_cell in range(2, (excel_activate.max_row)+1):
            if student_no ==  excel_activate['A'+str(each_cell)].value:
                Found = True
                break
            else:
                Found=False
        if(Found == True):
            Edit_form = Toplevel()
            Edit_form.geometry('500x1000')
            Edit_form.title('Edit Data from Excel')

            EditLabel = Label(Edit_form, text="Edit Form ", font=("Helvetica", 16))
            EditLabel.pack()

            student_noLbl=Label(Edit_form,text="Student No.",font=("bold",12),pady=(18))
            student_noLbl.pack()
            
            student_noExcel = StringVar()
            full_nameExcel = StringVar()
            emailExcel = StringVar()
            genderExcel = StringVar()
            courseExcel = StringVar()
            noExcel = StringVar()
            allowanceExcel = StringVar()
            addressExcel = StringVar()

            student_noTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=student_noExcel)
            student_noTxt.pack()

            student_noChoice = IntVar()
            student_noChk = Checkbutton(Edit_form, text="same as before", variable=student_noChoice, command=lambda:get_existing_student_no())
            student_noChk.pack()
            
            full_name=Label(Edit_form,text="Full Name",font=("bold",12),pady=(15))
            full_name.pack()

            full_nameTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=full_nameExcel)
            full_nameTxt.pack()

            full_nameChoice = IntVar()
            full_nameChk = Checkbutton(Edit_form, text="same as before", variable=full_nameChoice, command=lambda:get_existing_full_name())
            full_nameChk.pack()

            email=Label(Edit_form,text="Email",font=("bold",12),pady=(15))
            email.pack()

            emailTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=emailExcel)
            emailTxt.pack()

            emailChoice = IntVar()
            emailChk = Checkbutton(Edit_form, text="same as before", variable=emailChoice, command=lambda:get_existing_email())
            emailChk.pack()

            gender=Label(Edit_form,text="Gender",font=("bold",12),pady=(15))
            gender.pack()

            genderTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=genderExcel)
            genderTxt.pack()

            genderChoice = IntVar()
            genderChk = Checkbutton(Edit_form, text="same as before", variable=genderChoice, command=lambda:get_existing_gender())
            genderChk.pack()

            course=Label(Edit_form,text="Course",font=("bold",12),pady=(15))
            course.pack()

            courseTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=courseExcel)
            courseTxt.pack()

            courseChoice = IntVar()
            courseChk = Checkbutton(Edit_form, text="same as before", variable=courseChoice, command=lambda:get_existing_course())
            courseChk.pack()

            no=Label(Edit_form,text="Contact No.",font=("bold",12),pady=(15))
            no.pack()

            noTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=noExcel)
            noTxt.pack()

            noChoice = IntVar()
            noChk = Checkbutton(Edit_form, text="same as before", variable=noChoice, command=lambda:get_existing_no())
            noChk.pack()

            allowance=Label(Edit_form,text="Allowance",font=("bold",12),pady=(15))
            allowance.pack()

            allowanceTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=allowanceExcel)
            allowanceTxt.pack()

            allowanceChoice = IntVar()
            allowanceChk = Checkbutton(Edit_form, text="same as before", variable=allowanceChoice, command=lambda:get_existing_allowance())
            allowanceChk.pack()

            address=Label(Edit_form,text="Address",font=("bold",12),pady=(15))
            address.pack()

            addressTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=addressExcel)
            addressTxt.pack()

            addressChoice = IntVar()
            addressChk = Checkbutton(Edit_form, text="same as before", variable=addressChoice, command=lambda:get_existing_address())
            addressChk.pack()

            def get_existing_student_no():
                if student_noChoice.get()==1:
                    student_noOld = excel_activate['A'+str(each_cell)].value
                    student_noExcel.set(student_noOld)
                elif student_noChoice.get() ==0:
                    student_noExcel.set("")
            def get_existing_full_name():
                if full_nameChoice.get()==1:
                    full_nameOld = excel_activate['B'+str(each_cell)].value
                    full_nameExcel.set(full_nameOld)
                elif full_nameChoice.get() ==0:
                    full_nameExcel.set("")
            def get_existing_email():
                if emailChoice.get()==1:
                    emailOld = excel_activate['C'+str(each_cell)].value
                    emailExcel.set(emailOld)
                elif emailChoice.get() ==0:
                    emailExcel.set("")
            def get_existing_gender():
                if genderChoice.get()==1:
                    genderOld = excel_activate['D'+str(each_cell)].value
                    genderExcel.set(genderOld)
                elif genderChoice.get() ==0:
                    genderExcel.set("")
            def get_existing_course():
                if courseChoice.get()==1:
                    courseOld = excel_activate['E'+str(each_cell)].value
                    courseExcel.set(courseOld)
                elif courseChoice.get() ==0:
                    courseExcel.set("")
            def get_existing_no():
                if noChoice.get()==1:
                    noOld = excel_activate['F'+str(each_cell)].value
                    noExcel.set(noOld)
                elif noChoice.get() ==0:
                    noExcel.set("")
            def get_existing_allowance():
                if allowanceChoice.get()==1:
                    allowanceOld = excel_activate['G'+str(each_cell)].value
                    allowanceExcel.set(allowanceOld)
                elif allowanceChoice.get() ==0:
                    allowanceExcel.set("")
            def get_existing_address():
                if addressChoice.get()==1:
                    addressOld = excel_activate['H'+str(each_cell)].value
                    addressExcel.set(addressOld)
                elif addressChoice.get() ==0:
                    addressExcel.set("")
            
            def update():
                excel_activate['A'+str(each_cell)].value = student_noTxt.get()
                excel_activate['B'+str(each_cell)].value = full_nameTxt.get()
                excel_activate['C'+str(each_cell)].value = emailTxt.get()
                excel_activate['D'+str(each_cell)].value = genderTxt.get()
                excel_activate['E'+str(each_cell)].value = courseTxt.get()
                excel_activate['F'+str(each_cell)].value = noTxt.get()
                excel_activate['G'+str(each_cell)].value = allowanceTxt.get()
                excel_activate['H'+str(each_cell)].value = addressTxt.get()

                excel_con.save('Sample_file.xlsx')
                messagebox.showinfo("UPDATED","DATA HAS BEEN UPDATED")
                Edit_form.destroy()
                refresh_data(tv1)
            EditBtn = Button(Edit_form, width=15, font=("Arial", 20), text="Update Value",command=lambda:update(), bg="#b1f2ff")
            EditBtn.pack(padx=10, pady=9)

            Edit_form.mainloop()

    topframe = Frame(NewRoot, height=80, width=1205, bg="#CCAAFF")
    topframe.grid(row=0, column=0, columnspan=6, rowspan=2)
    leftframe = Frame(NewRoot, height=600, width=600, bd=3, relief="solid", bg="#CCAAFF")
    leftframe.grid(row=2, column=0, columnspan=3, rowspan=6, sticky=W)
    rightframe = Frame(NewRoot, height=600, width=600, bd=3, relief="solid" ,bg="#CCAAFF")
    rightframe.grid(row=2, column=3, columnspan=3, rowspan=6, sticky=W)
    bottomframe = Frame(NewRoot, height=80, width=1205, bg="#CCAAFF")
    bottomframe.grid(row=8, column=0, columnspan=6, rowspan=2, sticky=W)
    
    # topframe
    sts = Label(topframe, text="STUDENT RECORDS SYSTEM", font=("", 40), bg="#CCAAFF", width=40)
    sts.grid(row=0, column=0)

    # leftframe
    path = "S.png"
    photo0 = ImageTk.PhotoImage(file = path)
    pic = Label(leftframe, image= photo0,height=600, width=700)
    pic.place(x=0, y=0)
    student_no = Label(leftframe, text="Student No.", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    studentno_Entry = Entry(leftframe, width=25, font=("", 23))
    student_no.grid(row=0, column=0)
    studentno_Entry.grid(row=0, column=1, padx=10, pady=10,)

    full_name = Label(leftframe, text="Full Name", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    fullname_Entry = Entry(leftframe, width=25, font=("", 23))
    full_name.grid(row=1, column=0)
    fullname_Entry.grid(row=1, column=1, padx=10, pady=10)

    email = Label(leftframe, text="Email", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    email_Entry = Entry(leftframe, width=25, font=("", 23))
    email.grid(row=2, column=0)
    email_Entry.grid(row=2, column=1, padx=10, pady=10)

    gender = Label(leftframe, text="Gender", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    gender_Entry = Entry(leftframe, width=25, font=("", 23))
    gender.grid(row=3, column=0)
    gender_Entry.grid(row=3, column=1, padx=10, pady=10)

    course = Label(leftframe, text="Course", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    course_Entry = Entry(leftframe, width=25, font=("", 23))
    course.grid(row=4, column=0)
    course_Entry.grid(row=4, column=1, padx=10, pady=10)

    no = Label(leftframe, text="Contact No.", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    no_Entry = Entry(leftframe, width=25, font=("", 23))
    no.grid(row=5, column=0)
    no_Entry.grid(row=5, column=1, padx=10, pady=10)

    allowance = Label(leftframe, text="Allowance", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    allowance_Entry = Entry(leftframe, width=25, font=("", 23))
    allowance.grid(row=6, column=0)
    allowance_Entry.grid(row=6, column=1, padx=10, pady=10)

    address = Label(leftframe, text="Address", width=12, pady=20, font=("", 18), bg="#CCAAFF")
    address_Entry = Entry(leftframe, width=25, font=("", 23))
    address.grid(row=7, column=0)
    address_Entry.grid(row=7, column=1, padx=10, pady=10)

    # rightframe
    student_details = Label(rightframe, text="Student Details", width=31, pady=25, font=("", 18), bg="#CCAAFF")
    student_detailsT = Text(rightframe, width=44, font=("", 18), height=18)
    student_details.grid(row=1, column=1, columnspan=2)
    student_detailsT.grid(row=2, column=1)

    # bottomframe
    add_btn = Button(bottomframe, text="Add Record", command=add_record, width=15, pady=5, font=("", 20), bg="#CCBBFF")
    add_btn.grid(row=0, column=0)

    save_btn = Button(bottomframe, text="Save", command=save_records, width=15, pady=5, font=("", 20), bg="#CCBBFF")
    save_btn.grid(row=0, column=1)

    print_btn = Button(bottomframe, text="Print", command=print_records, width=15, pady=5, font=("", 20), bg="#CCBBFF")
    print_btn.grid(row=0, column=2, rowspan=2)

    reset_btn = Button(bottomframe, text="Reset", command=reset_fields, width=15, pady=5, font=("", 20), bg="#CCBBFF")
    reset_btn.grid(row=0, column=3)   

    exit_btn = Button(bottomframe, text="Exit", command=exit_interface, width=15, pady=5, font=("", 20), bg="#CCBBFF")
    exit_btn.grid(row=0, column=4)

    search_btn = Button(bottomframe, text="Search", command=search_data, width=15, pady=5, font=("", 20), bg="#CCBBFF")
    search_btn.grid(row=1, column=0)

    view_btn=Button(bottomframe,text="View Data", command=view_data ,width=15, pady=5, font=("", 20), bg="#CCBBFF")
    view_btn.grid(row=1, column=1)

    del_btn=Button(bottomframe,text="Delete", command=delete_data ,width=15, pady=5, font=("", 20), bg="#CCBBFF")
    del_btn.grid(row=1, column=3)

    edit_btn = Button(bottomframe, text="Edit", command=edit_data ,width=15, pady=5, font=("", 20), bg="#CCBBFF")
    edit_btn.grid(row=1, column=4)

    NewRoot.mainloop()

root = tk.Tk()
root.geometry("475x265")
root.title("Login Interface") 
#root.resizable(False, False)

path = "pic.jpg"
bg = ImageTk.PhotoImage(file = path)
my_canva = Canvas(root, height=265, width=475)
my_canva.pack(fill="both")

my_canva.create_image(0,0,image=bg, anchor="nw")

my_canva.create_text(250, 20, text="Sign In", font=("Arial", 20),fill="black")

userLabel = LabelFrame(root, text="Username", width=15, bg="white")
username = Entry(userLabel,font=("Arial", 18))
username.pack()
username_w = my_canva.create_window(110, 60, anchor="nw", window=userLabel)

passwordLabel = LabelFrame(root, text="Password", width=15)
password = Entry(passwordLabel,font=("Arial", 18))
password.pack()
password_w = my_canva.create_window(110, 110, anchor="nw", window=passwordLabel)

button1 = Button(root, text="Login", width=20, bg="lightblue", command=lambda:login())
button1_window = my_canva.create_window(170, 170, anchor="nw", window=button1)

tk.mainloop()