from  tkinter import *
import tkinter as tk 
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk

def newinterface():
		NewRoot = Toplevel()
		NewRoot.geometry("1030x780")
		NewRoot.title("New Interface")
		NewRoot.configure(bg="black")

		#Excel File
		excel_con = Workbook()
		excel_con = load_workbook("Sample_file.xlsx")
		excel_data = excel_con['Data']

		#ADD RECORD BBUTTON FUNCTION
		def add_record():
				student_no = studentno_Entry.get()
				full_name = fullname_Entry.get()
				email = email_Entry.get()
				gender = gender_var.get()
				course = course_var.get()
				contact_no = no_Entry.get()
				address = address_Entry.get()
				
				Found = False
				for each_cell in range(2, excel_data.max_row + 1):
					if student_no == excel_data["A"+ str(each_cell)].value or full_name == excel_data["B"+ str(each_cell)].value:
						Found = True
						break
				if Found == True:
					messagebox.showerror("Data", "Data Already Exist")
				else:
					lastrow = str(excel_data.max_row + 1)
					excel_data["A"+lastrow] = student_no
					excel_data["B"+lastrow] = full_name
					excel_data["C"+lastrow] = email
					excel_data["D"+lastrow] = gender
					excel_data["E"+lastrow] = course
					excel_data["F"+lastrow] = contact_no
					excel_data["G"+lastrow] = address

					excel_con.save("Sample_file.xlsx")
					messagebox.showinfo("Save Records", "Records saved successfully!")
					refresh_data(tv1)

				record = f"Student Ref: {student_no}\n"
				record += f"Full Name: {full_name}\n"
				record += f"Email: {email}\n"
				record += f"Gender: {gender}\n"
				record += f"Course: {course}\n"
				record += f"Contact No.: {contact_no}\n"
				record += f"Address: {address}\n"
				
				student_detailsT.insert(END, record)

				studentno_Entry.delete(0, END)
				fullname_Entry.delete(0, END)
				email_Entry.delete(0, END)
				gender_var.set("none")
				course_var.set("")
				no_Entry.delete(0, END)
				address_Entry.delete(0, END)
		#REFRESH FUNCTION
		def refresh_data(tree):
				tree.delete(*tree.get_children())
				data = get_updated_data()
				for each_cell in range(2, (excel_data.max_row)+1):
					tv1.insert(parent='', index="end", text=str(each_cell),values=(excel_data['A'+str(each_cell)].value,excel_data['B'+str(each_cell)].value, excel_data['C'+str(each_cell)].value, excel_data['D'+str(each_cell)].value, excel_data['E'+str(each_cell)].value, excel_data['F'+str(each_cell)].value, excel_data['G'+str(each_cell)].value, excel_data['H'+str(each_cell)].value))

		#UPDATE FUNCTION
		def get_updated_data():
				updated_value = list()
				for each_cell in range(2, (excel_data.max_row)+1):     
					updated_value.append([excel_data['A'+str(each_cell)].value,excel_data['B'+str(each_cell)].value, excel_data['C'+str(each_cell)].value, excel_data['D'+str(each_cell)].value,excel_data['E'+str(each_cell)].value,excel_data['F'+str(each_cell)].value,excel_data['G'+str(each_cell)].value,excel_data['H'+str(each_cell)].value])
				return updated_value

		#EXIT BUTTON FUNCTION
		def exit_interface():
			NewRoot.destroy()

		#PRINT BUTTON FUNCTION
		def print_records():
			student_details_text = student_detailsT.get("1.0", END)
			messagebox.showinfo("Print Records", student_details_text)
			student_details_text = student_detailsT.delete("1.0", END)

		#RESET BUTTON FUNCTION
		def reset_fields():
				studentno_Entry.delete(0, END)
				fullname_Entry.delete(0, END)
				email_Entry.delete(0, END)
				gender_var.set("none")
				course_var.set("")
				no_Entry.delete(0, END)
				address_Entry.delete(0, END)
				search_box.delete(0, END)
				student_detailsT.delete("1.0", END)

		def search_data():
			data = []

			for i in excel_data.iter_rows(values_only=True):
				data.append(i)

				tv1.delete(*tv1.get_children())
				lists = []
				for i in data:
					x = False
					for j in i:
						if j != None:
							if search_box.get().lower() in j.lower():
								x = True
								break
					if x:
						lists.append(i)
				
				for i in lists:
					tv1.insert('', index=END, values=i)

		def delete_data():
			student_no = search_box.get()
			pos = 1
			Found = False
			for each_cell in excel_data.iter_rows(values_only=True):
				if student_no == each_cell[0]:
					Found = True
					break
				pos += 1
			if(Found == True):
				excel_data.delete_rows(pos)
				messagebox.showinfo("INFO","DATA DELETED")
				clear_entries()
			excel_con.save('Sample_file.xlsx')
			refresh_data(tv1)

		def clear_entries():
			studentno_Entry.delete(0, END)
			fullname_Entry.delete(0, END)
			email_Entry.delete(0, END)
			gender_var.set("none")
			course_var.set("")
			no_Entry.delete(0, END)
			search_box.delete(0, END)
			address_Entry.delete(0, END)

		def edit_data():
			search_bo = search_box.get()
			for each_cell in range(2, (excel_data.max_row)+1):
				if search_bo ==  excel_data['A'+str(each_cell)].value:
					Found = True
					break
				else:
					Found=False
			if(Found == True):
				Edit_form = Toplevel()
				Edit_form.geometry('300x800')
				Edit_form.title('Edit Data from Excel')

				EditLabel = Label(Edit_form, text="Edit Form ", font=("Helvetica", 16))
				EditLabel.pack()

				student_noLbl=Label(Edit_form,text="Student No.",font=("bold",12),pady=(18))
				student_noLbl.pack()
				
				student_noExcel = StringVar()
				full_nameExcel = StringVar()
				emailExcel = StringVar()
				genderExcel = StringVar()
				courseExcel = StringVar()
				noExcel = StringVar()
				addressExcel = StringVar()

				student_noTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=student_noExcel)
				student_noTxt.pack()

				student_noChoice = IntVar()
				student_noChk = Checkbutton(Edit_form, text="same as before", variable=student_noChoice, command=lambda:get_existing_student_no())
				student_noChk.pack()
				
				full_name=Label(Edit_form,text="Full Name",font=("bold",12),pady=(15))
				full_name.pack()

				full_nameTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=full_nameExcel)
				full_nameTxt.pack()

				full_nameChoice = IntVar()
				full_nameChk = Checkbutton(Edit_form, text="same as before", variable=full_nameChoice, command=lambda:get_existing_full_name())
				full_nameChk.pack()

				email=Label(Edit_form,text="Email",font=("bold",12),pady=(15))
				email.pack()

				emailTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=emailExcel)
				emailTxt.pack()

				emailChoice = IntVar()
				emailChk = Checkbutton(Edit_form, text="same as before", variable=emailChoice, command=lambda:get_existing_email())
				emailChk.pack()

				gender=Label(Edit_form,text="Gender",font=("bold",12),pady=(15))
				gender.pack()

				genderTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=genderExcel)
				genderTxt.pack()

				genderChoice = IntVar()
				genderChk = Checkbutton(Edit_form, text="same as before", variable=genderChoice, command=lambda:get_existing_gender())
				genderChk.pack()

				course=Label(Edit_form,text="Course",font=("bold",12),pady=(15))
				course.pack()

				courseTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=courseExcel)
				courseTxt.pack()

				courseChoice = IntVar()
				courseChk = Checkbutton(Edit_form, text="same as before", variable=courseChoice, command=lambda:get_existing_course())
				courseChk.pack()

				no=Label(Edit_form,text="Contact No.",font=("bold",12),pady=(15))
				no.pack()

				noTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=noExcel)
				noTxt.pack()

				noChoice = IntVar()
				noChk = Checkbutton(Edit_form, text="same as before", variable=noChoice, command=lambda:get_existing_no())
				noChk.pack()

				address=Label(Edit_form,text="Address",font=("bold",12),pady=(15))
				address.pack()

				addressTxt=Entry(Edit_form, width=25, font=('Helvetica',12),textvariable=addressExcel)
				addressTxt.pack()

				addressChoice = IntVar()
				addressChk = Checkbutton(Edit_form, text="same as before", variable=addressChoice, command=lambda:get_existing_address())
				addressChk.pack()

				def get_existing_student_no():
					if student_noChoice.get()==1:
						student_noOld = excel_data['A'+str(each_cell)].value
						student_noExcel.set(student_noOld)
					elif student_noChoice.get() ==0:
						student_noExcel.set("")
				def get_existing_full_name():
					if full_nameChoice.get()==1:
						full_nameOld = excel_data['B'+str(each_cell)].value
						full_nameExcel.set(full_nameOld)
					elif full_nameChoice.get() ==0:
						full_nameExcel.set("")
				def get_existing_email():
					if emailChoice.get()==1:
						emailOld = excel_data['C'+str(each_cell)].value
						emailExcel.set(emailOld)
					elif emailChoice.get() ==0:
						emailExcel.set("")
				def get_existing_gender():
					if genderChoice.get()==1:
						genderOld = excel_data['D'+str(each_cell)].value
						genderExcel.set(genderOld)
					elif genderChoice.get() ==0:
						genderExcel.set("")
				def get_existing_course():
					if courseChoice.get()==1:
						courseOld = excel_data['E'+str(each_cell)].value
						courseExcel.set(courseOld)
					elif courseChoice.get() ==0:
						courseExcel.set("")
				def get_existing_no():
					if noChoice.get()==1:
						noOld = excel_data['F'+str(each_cell)].value
						noExcel.set(noOld)
					elif noChoice.get() ==0:
						noExcel.set("")
				def get_existing_address():
					if addressChoice.get()==1:
						addressOld = excel_data['H'+str(each_cell)].value
						addressExcel.set(addressOld)
					elif addressChoice.get() ==0:
						addressExcel.set("")
				
				def update():
					excel_data['A'+str(each_cell)].value = student_noTxt.get()
					excel_data['B'+str(each_cell)].value = full_nameTxt.get()
					excel_data['C'+str(each_cell)].value = emailTxt.get()
					excel_data['D'+str(each_cell)].value = genderTxt.get()
					excel_data['E'+str(each_cell)].value = courseTxt.get()
					excel_data['F'+str(each_cell)].value = noTxt.get()
					excel_data['G'+str(each_cell)].value = addressTxt.get()

					excel_con.save('Sample_file.xlsx')
					messagebox.showinfo("UPDATED","DATA HAS BEEN UPDATED")
					Edit_form.destroy()
					search_box.delete(0, END)
					refresh_data(tv1)
				EditBtn = Button(Edit_form, width=15, font=("Arial", 12), text="Update Value",command=lambda:update(), bg="#b1f2ff")
				EditBtn.pack(padx=10, pady=9)

				Edit_form.mainloop()

		#Frames
		topframe = Frame(NewRoot, height=70, width=1030, bg="#CCAAFF")
		topframe.grid(row=0, column=0, columnspan=6, rowspan=2, sticky=W)
		centerframe = Frame(NewRoot, height=590, width=1030, bd=3)
		centerframe.grid(row=2, column=0, columnspan=3, rowspan=6, sticky=W)
		bottomframe = Frame(NewRoot, height=100, width=1030, bg="#CCAAFF")
		bottomframe.grid(row=8, column=0, columnspan=6, rowspan=2, sticky=W)

		# topframe
		pat = "gr.jpg"
		b = Image.open(pat)
		resize_b = b.resize((1030, 100))
		b = ImageTk.PhotoImage(resize_b)
		my_canv = Canvas(topframe, height=70, width=1030)
		my_canv.grid(row=0, column=0)
		my_canv.create_image(0,0,image=b, anchor="nw")
		my_canv.create_text(515, 40, text="STUDENT RECORDS SYSTEM", font=("System", 41),fill="black")

		#center
		path = "gr.jpg"
		bg = Image.open(path)
		resize_bg = bg.resize((1030, 590))
		bg = ImageTk.PhotoImage(resize_bg)
		my_canva = Canvas(centerframe, height=590, width=1030)
		my_canva.grid()
		my_canva.create_image(0,0,image=bg, anchor="nw")

		student_no = LabelFrame(centerframe, text="Student No.", font=("Arial", 10))
		studentno_Entry = Entry(student_no, width=30, font=("Arial", 15))
		student_no.grid(row=0, column=0)
		studentno_Entry.grid(row=0, column=1, padx=10, pady=10,)
		student_no_canva = my_canva.create_window(50, 30, anchor="nw", window=student_no)

		full_name = LabelFrame(centerframe, text="Full Name", font=("Arial", 10))
		fullname_Entry = Entry(full_name, width=30, font=("Arial", 15))
		full_name.grid(row=1, column=0)
		fullname_Entry.grid(row=1, column=1, padx=10, pady=10)
		full_name_canva = my_canva.create_window(50, 110, anchor="nw", window=full_name)

		email = LabelFrame(centerframe, text="Email", font=("Arial", 10))
		email_Entry = Entry(email, width=30, font=("Arial", 15))
		email.grid(row=2, column=0)
		email_Entry.grid(row=2, column=1, padx=10, pady=10)
		email_canva = my_canva.create_window(50, 190, anchor="nw", window=email)

		gender = LabelFrame(centerframe, text="Gender", font=("Arial", 10))
		gender_var = StringVar()
		gender_var.set("none")
		male_R = Radiobutton(gender, text="Male", variable=gender_var, value="Male", width=17, font=("Arial", 12))
		female_R = Radiobutton(gender, text="Female", variable=gender_var, value="Female", width=16, font=("Arial", 12))
		male_R.grid(row=3, column=0)
		female_R.grid(row=3, column=1)
		gender_canva = my_canva.create_window(50, 270, anchor="nw", window=gender)

		course = LabelFrame(centerframe, text="Course", font=("Arial", 10))
		course_var = StringVar()
		course_list = ["BSIT", "BSA", "BSAIS", "ABELS", "BSSW", "BSE", "DHRS", "BSPA", "BTVTE"]
		course_combo = ttk.Combobox(course, values=course_list, textvariable=course_var, font=("Arial", 15),width=30)
		course_combo.grid(row=4, column=0, padx=1, pady=10)
		course_canva = my_canva.create_window(50, 330, anchor="nw", window=course)

		no = LabelFrame(centerframe, text="Contact No.", font=("Arial", 10))
		no_Entry = Entry(no, width=30, font=("Arial", 15))
		no.grid(row=5, column=0)
		no_Entry.grid(row=5, column=1, padx=10, pady=10)
		no_canva = my_canva.create_window(50, 410, anchor="nw", window=no)

		address = LabelFrame(centerframe, text="Address", font=("Arial", 10))
		address_Entry = Entry(address, width=30, font=("Arial", 15))
		address.grid(row=6, column=0)
		address_Entry.grid(row=7, column=1, padx=10, pady=10)
		address_canva = my_canva.create_window(50, 490, anchor="nw", window=address)

		student_details = LabelFrame(centerframe, text="Student Details", font=("Arial", 15))
		student_detailsT = Text(student_details, width=51, font=("Arial", 15), height=8)
		student_details.grid(row=1, column=1, columnspan=2)
		student_detailsT.grid(row=2, column=1)
		student_details_canva = my_canva.create_window(420, 30, anchor="nw", window=student_details)

		firstclick = True
		def on_search_box_click(event):     
			global firstclick

			if firstclick: 
				firstclick = False
				search_box.delete(0, "end")

		search_boxf = LabelFrame(centerframe, text="Search, Delete, Edit", font=("Arial", 12))
		search_box = Entry(search_boxf, width=39, font=("Arial", 12), bg=search_boxf['bg'], bd=0)
		search_box.insert(0, "Please input what you what to search, delete or edit")
		search_box.bind('<FocusIn>', on_search_box_click)
		search_boxf.grid(row=1, column=0)
		search_box.grid(row=1, column=1, padx=10, pady=10)



		s = Button(search_boxf, text="Search", bd=0, font=("Arial", 12),width=7, command=lambda:search_data())
		s.grid(row=1, column=2)
		d = Button(search_boxf, text="Delete", bd=0, font=("Arial", 12),width=7, command=lambda:delete_data())
		d.grid(row=1, column=3)
		e = Button(search_boxf, text="Edit", bd=0, font=("Arial", 12), width=5, command=lambda:edit_data())
		e.grid(row=1, column=4)
		search_boxf_canva = my_canva.create_window(420, 250, anchor="nw", window=search_boxf)

		t = LabelFrame(centerframe, text="VIEW", font=("Arial", 15))
		global tv1
		tv1 = ttk.Treeview(t, show='headings')
		treescrolly = Scrollbar(t, orient="vertical", command=tv1.yview)
		treescrollx = Scrollbar(t, orient="horizontal", command=tv1.xview)
		tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
		treescrollx.pack(side ="bottom",fill ="x")
		treescrolly.pack(side ="right",fill="y")  

		tv1['columns'] = ("Student No.", "Fullname", "Email", "Gender", "Course", "Contact No.", "Address")
		tv1.column("#0", width=120, minwidth=25)
		tv1.column("Student No.", anchor=W, width=120)
		tv1.column("Fullname",  anchor=W, width=120)
		tv1.column("Email", anchor=W, width=120)
		tv1.column("Gender", anchor=W, width=120)
		tv1.column("Course", anchor=W, width=120)
		tv1.column("Contact No.", anchor=W, width=120)
		tv1.column("Address", anchor=W, width=120)

		tv1.heading("#0", text="Label", anchor=W)
		tv1.heading("Student No.", text="Student No.", anchor=W)
		tv1.heading("Fullname", text="Fullname", anchor=W)
		tv1.heading("Email", text="Email", anchor=W)
		tv1.heading("Gender", text="Gender", anchor=W)
		tv1.heading("Course", text="Course", anchor=W)
		tv1.heading("Contact No.", text="Contact No.", anchor=W)
		tv1.heading("Address", text="Address", anchor=W)

		for each_cell in range(2, (excel_data.max_row)+1):
			tv1.insert(parent='', index="end", text=str(each_cell),values=(excel_data['A'+str(each_cell)].value,excel_data['B'+str(each_cell)].value, excel_data['C'+str(each_cell)].value, excel_data['D'+str(each_cell)].value, excel_data['E'+str(each_cell)].value, excel_data['F'+str(each_cell)].value, excel_data['G'+str(each_cell)].value, excel_data['H'+str(each_cell)].value))
		tv1.pack()
		t.place(x=420, y=315, width=565, height=245)

		# bottomframe
		pa = "gr.jpg"
		c = Image.open(pa)
		resize_c = c.resize((1030, 100))
		c = ImageTk.PhotoImage(resize_b)
		my_can = Canvas(bottomframe, height=100, width=1030)
		my_can.grid(row=0, column=0)
		my_can.create_image(0,0,image=c, anchor="nw")


		#BUTTONS
		add_btn = Button(bottomframe, text="Add Record", font=("System", 25), width=13, command=lambda:add_record())
		add_btn_w = my_can.create_window(10, 25, anchor="nw", window=add_btn)

		print_btn = Button(bottomframe, text="Print", font=("System", 25), width=13, command=lambda:print_records())
		print_btn_w = my_can.create_window(270, 25, anchor="nw", window=print_btn)

		reset_btn = Button(bottomframe, text="Reset", font=("System", 25), width=13, command=lambda:reset_fields())
		reset_btn_w = my_can.create_window(530, 25, anchor="nw", window=reset_btn)   

		exit_btn = Button(bottomframe, text="Exit", font=("System", 25), width=13, command=lambda:exit_interface())
		exit_btn_w = my_can.create_window(790, 25, anchor="nw", window=exit_btn)


		NewRoot.mainloop()

#FIRST INTERFACE
root = tk.Tk()
root.geometry("1120x600")
root.title("Login Interface")
root.resizable(False, False)

excel_con = Workbook()
excel_con = load_workbook("Sample_file.xlsx")
excel_data = excel_con['Data']
excel_old = excel_con["Old Students"]
excel_new = excel_con["New Students"]

def oldReg():
	newroot = Toplevel()
	newroot.geometry("600x500")
	newroot.title("Register Interface For Old Students") 
	newroot.resizable(False, False)

	path = "l.webp"
	bg = Image.open(path)
	resize_bg = bg.resize((600, 500))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(newroot, height=500, width=600)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(270, 70, text="Register For Old Students", font=("Arial", 20),fill="black")

	studentnol = LabelFrame(newroot, text="Student No.", width=15, bg="white")
	studentno = Entry(studentnol,font=("Arial", 18))
	studentno.pack()
	studentno_w = my_canva.create_window(130, 100, anchor="nw", window=studentnol)

	year_sectionl = LabelFrame(newroot, text="Year And Section", width=15, bg="white")
	year_section = Entry(year_sectionl,font=("Arial", 18))
	year_section.pack()
	year_section_w = my_canva.create_window(130, 170, anchor="nw", window=year_sectionl)

	passwordl = LabelFrame(newroot, text="Password", width=15, bg="white")
	password = Entry(passwordl,font=("Arial", 18))
	password.pack()
	password_w = my_canva.create_window(130, 240, anchor="nw", window=passwordl)

	def oldlyreg():
		excel_old.append((studentno.get(), password.get(), year_section.get()))
		messagebox.showinfo("Register Message", "Register Successful")
		excel_con.save("Sample_file.xlsx")

	r_button = Button(newroot, text="Register", width=20, bg="lightblue", font=("Arial", 12) , command=lambda:oldlyreg())
	button = my_canva.create_window(170, 310, anchor="nw", window=r_button)

	newroot.mainloop()
	

def newReg():
	
	nroot = Toplevel()
	nroot.geometry("600x500")
	nroot.title("Register Interface For New Students") 
	nroot.resizable(False, False)

	path = "r.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((600, 500))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(nroot, height=500, width=600)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(260, 60, text="Register For New Students", font=("Arial", 20),fill="black")

	fullnl = LabelFrame(nroot, text="Fullname", width=15, bg="white")
	fulln = Entry(fullnl,font=("Arial", 18))
	fulln.pack()
	fulln_w = my_canva.create_window(130, 90, anchor="nw", window=fullnl)

	year_secl = LabelFrame(nroot, text="Year And Section", width=15, bg="white")
	year_sec = Entry(year_secl,font=("Arial", 18))
	year_sec.pack()
	year_sec_w = my_canva.create_window(130, 160, anchor="nw", window=year_secl)

	passwl = LabelFrame(nroot, text="Password", width=15, bg="white")
	passw = Entry(passwl,font=("Arial", 18))
	passw.pack()
	passw_w = my_canva.create_window(130, 230, anchor="nw", window=passwl)

	def newlyreg():
		excel_new.append((fulln.get(), passw.get(), year_sec.get()))
		messagebox.showinfo("Register Message", "Register Successful")
		excel_con.save("Sample_file.xlsx")

	r_button1 = Button(nroot, text="Register", width=20, bg="lightblue", font=("Arial",12),command=lambda:newlyreg())
	r_button = my_canva.create_window(170, 300, anchor="nw", window=r_button1)

	nroot.mainloop()

#Register Function
def reg():
	if str_var.get() == "Old":
		oldReg()
	elif str_var.get() == "New":
		newReg()
	else:
		messagebox.showerror("REGISTER ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")
		
def newLog():
	groot = Toplevel()
	groot.geometry("700x350")
	groot.title("Login Interface For New Students") 
	#groot.resizable(False, False)

	path = "login.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((700, 350))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(groot, height=350, width=700)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(450, 50, text="Sign In For New Students", font=("Arial", 20),fill="black")

	fullnamel = LabelFrame(groot, text="Fullname", width=15, bg="white")
	fullname = Entry(fullnamel,font=("Arial", 18))
	fullname.pack()
	fullname_w = my_canva.create_window(310, 90, anchor="nw", window=fullnamel)

	passwordLabel = LabelFrame(groot, text="Password", width=15, bg="white")
	passwor = Entry(passwordLabel,font=("Arial", 18), show="*")
	passwor.pack()
	password = my_canva.create_window(310, 170, anchor="nw", window=passwordLabel)

	def getUsers(fullname, passwor):
		id = 1
		isExisted = False
		for data in excel_new.iter_rows(values_only=True):
			if data[0] == fullname and data[1] == passwor:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			groot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			groot.destroy()
			newinterface()

	button1 = Button(groot, text="Login", width=20, bg="lightblue", font=("Arial", 12), command=lambda:getUsers(fullname.get(), passwor.get(), newinterface()))
	button1 = my_canva.create_window(350, 250, anchor="nw", window=button1)
	groot.mainloop()	


def oldLog():
	vroot = Toplevel()
	vroot.geometry("700x350")
	vroot.title("Login Interface For Old Students") 
	#vroot.resizable(False, False)

	path = "login.jpg"
	bg = Image.open(path)
	resize_bg = bg.resize((700, 350))
	bg = ImageTk.PhotoImage(resize_bg)
	my_canva = Canvas(vroot, height=350, width=700)
	my_canva.pack(fill="both")
	my_canva.create_image(0,0,image=bg, anchor="nw")

	my_canva.create_text(450, 50, text="Sign In For Old Students", font=("Arial", 20),fill="black")

	studentnl = LabelFrame(vroot, text="Student No", width=15, bg="white")
	studentn = Entry(studentnl,font=("Arial", 18))
	studentn.pack()
	studentn_w = my_canva.create_window(310, 90, anchor="nw", window=studentnl)

	passwol = LabelFrame(vroot, text="Password", width=15, bg="white")
	passwo = Entry(passwol,font=("Arial", 18), show="*")
	passwo.pack()
	passwo_w = my_canva.create_window(310, 170, anchor="nw", window=passwol)

	def getUser(studentn, passwo):
		id = 1
		isExisted = False
		for data in excel_old.iter_rows(values_only=True):
			if data[0] == studentn and data[1] == passwo:
				isExisted = True
				break
			id += 1
		if not isExisted:
			messagebox.showerror("Error", "Account not found\nPlease Register First")
			vroot.destroy()
		else:
			messagebox.showinfo("Login", "Login Successfuly")
			vroot.destroy()
			newinterface()

	button1 = Button(vroot, text="Login", width=20, bg="lightblue", font=("Arial",12), command=lambda:getUser(studentn.get(), passwo.get(),newinterface()))
	button1 = my_canva.create_window(350, 250, anchor="nw", window=button1)

	vroot.mainloop()
#Login function	
def log():
	if str_var.get() == "Old":
		oldLog()
	elif str_var.get() == "New":
		newLog()
	else:
		messagebox.showerror("LOGIN ERROR", "Pumili ka na dun sa dalwa kahit wag na ako")


#FRAMES
topframe = Frame(root, height=50, width=1120, bg="#023047")
topframe.pack(fill="x", side="top")
leftframe = Frame(root, height=600, width=350, bg="#023047", bd=0)
leftframe.pack(fill="y",side='left')
rigthframe = Frame(root, height=600, width=820)
rigthframe.pack(fill="both", side='right')

#TEXT IN TOPFRAME
l = Label(topframe,text="Dalubhasaan Ng Lungsod Ng Lucena", bg="#023047", fg="#FFB703", font=("Arial", 30), justify="center")
l.pack()

#LOGO IMAGE
b = (Image.open("dll_logo.png"))
resized_image = b.resize((190, 190))
new_image = ImageTk.PhotoImage(resized_image)
logo = Label(leftframe, image=new_image, bg="#023047")
logo.pack()

#RadioButton
rFrame = Frame(leftframe)

str_var = StringVar()
str_var.set("none")
old_R = Radiobutton(rFrame, text="Old", variable=str_var, value="Old", font=("Arial", 15),bg="#023047",fg="#FFB703")
new_R = Radiobutton(rFrame, text="New", variable=str_var, value="New", font=("Arial", 15),bg="#023047",fg="#FFB703")
old_R.grid(row=0, column=0)
new_R.grid(row=0, column=1)

rFrame.pack()

#Buttons
rb = Button(leftframe, text="Register", width=15, font=("Arial", 12), command=lambda: reg())
lb = Button(leftframe, text="Login", width=15, font=("Arial", 12), command=lambda: log())
rb.pack(pady=15)
lb.pack()

def animateAbout():
	global aboutIterate, txt
	if aboutIterate < len(a):
		txt += a[aboutIterate] + "\n"
		a_l1.config(text=txt)
		aboutIterate += 1
		a_l1.after(200, animateAbout)

#TEXT BUTTON 
def showProf():
	prof.pack(fill="both")
	abt.pack_forget()

def showAbt():
	abt.pack(fill="both")
	prof.pack_forget()
	animateAbout()

abt_b = Button(topframe, text="About", command=showAbt, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
abt_b.pack(side=RIGHT, anchor=N)
prof_b = Button(topframe, text="Profile", command=showProf, font=("Arial",15), bg="#023047", fg="#FFB703", bd=0)
prof_b.pack(side=RIGHT, anchor=N)

#FRAME FOR THE PROFILE AND ABOUT
prof = Frame(rigthframe, bg="#023047")

#SCHOOL IMAGE
bg = (Image.open("school.jpg"))
resized_imag = bg.resize((1120, 600))
new_imag = ImageTk.PhotoImage(resized_imag)
my_canva = Canvas(prof, height=600, width=1080)
my_canva.pack()
my_canva.create_image(0,0,image=new_imag, anchor="nw")
prof.pack()

aboutIterate = 0
txt = ""

#About
_a = "Dalubhasaan ng Lungsod ng Lucena"
a = [
	"One of the high impact programs of Mayor Roderick A. Alcala is free quality tertiary education.",
	"When he assumed office in 2012, Dalubhasaan ng Lungsod ng Lucena (DLL) was his vision of providing access to college education for free. \nMayor Alcala envisions DLL as an institution that would provide easy access to higher education and \nultimately develop the competencies of the youth of the city to meet the demands of the local industries and businesses.",
	"Through DLL, students from low-income families are able to enrol in degree programs at no cost. \nThe annual appropriation of the local government has allowed DLL to cover all its operation expenses \nincluding tuition and miscellaneous fees of students. \nThe college, operated, managed, and fully-subsidized by the City Government, implements a zero-collection policy.",
	"At present, DLL has a total of nine degree programs and continue to apply for additional academic programs to accommodate more scholars:",
	"Bachelor of Arts in Information Technology",
	"Bachelor of Arts in Public Administration",
	"Bachelor of Science in Accountancy",
	"Bachelor of Science in Accounting Information System",
	"Bachelor of Science in English Language Studies",
	"Bachelor of Science in Entrepreneurship",
	"Bachelor of Science in Social Work",
	"Bachelor in Technical Vocational Teacher’s Education",
	"Diploma in Hotel and Restaurant Services",
	"At present the former annex building of Lucena City Hall is being renovated to accommodate the growing student population of DLL."
]

abt = Frame(rigthframe, width=600, height=750)
a_l = Label(abt,text=_a,font=("System" ,25),fg="black",justify="center")
a_l1 = Label(abt,font=("System" ,16),fg="black",justify="left")

a_l.pack()
a_l1.pack(pady=20)


abt.pack()
abt.mainloop()

showProf()


root.mainloop()

